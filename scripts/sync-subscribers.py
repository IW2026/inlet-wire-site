#!/usr/bin/env python3
"""
Inlet Wire — Newsletter Subscriber Sync
---------------------------------------
Fetches signups from the Google Apps Script endpoint and saves to
a local Excel file at: <workspace>/subscribers.xlsx

Usage:
  python3 scripts/sync-subscribers.py

No config file needed — the Apps Script URL is built into this script.
"""

import json
import sys
import urllib.request
import urllib.error
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("ERROR: openpyxl not installed.")
    print("Run:  pip install openpyxl --break-system-packages")
    sys.exit(1)

# ── Config ────────────────────────────────────────────────────────────────────
APPS_SCRIPT_URL = (
    "https://script.google.com/macros/s/"
    "AKfycbw68p65bWFxFYZ5Y-vyErqeQI5acqXvqzm8VYeZ-HWQ8Icnp9EEnm53ZzFEmSCtdeLy"
    "/exec"
)

SCRIPT_DIR  = Path(__file__).resolve().parent   # site/scripts/
WORKSPACE   = SCRIPT_DIR.parent.parent          # Inlet Wire/
OUTPUT_FILE = WORKSPACE / "subscribers.xlsx"

# ── Fetch from Apps Script ────────────────────────────────────────────────────
def fetch_subscribers():
    url = APPS_SCRIPT_URL + "?action=export"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "InletWireSync/1.0"})
        with urllib.request.urlopen(req, timeout=20) as resp:
            data = json.loads(resp.read().decode())
            return data.get("subscribers", [])
    except urllib.error.HTTPError as e:
        print(f"ERROR fetching subscribers (HTTP {e.code}): {e.read().decode()}")
        sys.exit(1)
    except Exception as e:
        print(f"ERROR: {e}")
        sys.exit(1)

# ── Excel helpers ─────────────────────────────────────────────────────────────
HEADERS = ["email", "date_added", "source", "synced_at"]

def load_existing():
    if not OUTPUT_FILE.exists():
        return [], set()
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb.active
    rows, emails = [], set()
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        if row[0]:
            rows.append(dict(zip(HEADERS, row)))
            emails.add(str(row[0]).lower().strip())
    return rows, emails

def save_workbook(all_rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Subscribers"

    ws.append(HEADERS)
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="0A0A0A")
        cell.alignment = Alignment(horizontal="left")

    for r in all_rows:
        ws.append([r.get(h, "") for h in HEADERS])

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 22

    wb.save(OUTPUT_FILE)

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("Fetching subscribers from Apps Script...")
    remote = fetch_subscribers()
    print(f"  Found {len(remote)} subscriber(s) in Google Sheet.")

    existing_rows, existing_emails = load_existing()

    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    new_rows = []
    for r in remote:
        email = r.get("email", "").strip().lower()
        if email and email not in existing_emails:
            new_rows.append({
                "email":      email,
                "date_added": r.get("date_added", ""),
                "source":     r.get("source", "website"),
                "synced_at":  now_str,
            })

    if not new_rows:
        print(f"No new subscribers. Total on file: {len(existing_rows)}")
        return

    all_rows = existing_rows + new_rows
    save_workbook(all_rows)
    print(f"✓ Added {len(new_rows)} new subscriber(s). Total: {len(all_rows)}")
    print(f"  Saved to: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
